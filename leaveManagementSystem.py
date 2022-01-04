# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *
 
# globally declare wb and sheet variable
 
# opening the existing excel file
wb = load_workbook('D:\\5th Sem Online Classes\\ADP\\ADP Activity\\excel.xlsx')
 
# create the sheet object
sheet = wb.active
 
 
def excel():
     
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 30
    sheet.column_dimensions['H'].width = 30
    sheet.column_dimensions['I'].width = 30
    sheet.column_dimensions['J'].width = 60
 
    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "USN"
    sheet.cell(row=1, column=2).value = "Name"
    sheet.cell(row=1, column=3).value = "Branch"
    sheet.cell(row=1, column=4).value = "Semester"
    sheet.cell(row=1, column=5).value = "Section"
    sheet.cell(row=1, column=6).value = "Email ID"
    sheet.cell(row=1, column=7).value = "Phone No"
    sheet.cell(row=1, column=8).value = "Start Date"
    sheet.cell(row=1, column=9).value = "End Date"
    sheet.cell(row=1, column=10).value = "Leave Reason"
 
 
# Function to set focus (cursor)
def focus1(event):
    # set focus on the name_field box
    name_field.focus_set()
 
 
# Function to set focus
def focus2(event):
    # set focus on the branch_field box
    branch_field.focus_set()
 
 
# Function to set focus
def focus3(event):
    # set focus on the semester_field box
    semester_field.focus_set()
 
 
# Function to set focus
def focus4(event):
    # set focus on the section_field box
    section_field.focus_set()
 
 
# Function to set focus
def focus5(event):
    # set focus on the email_id_field box
    email_id_field.focus_set()
 
 
# Function to set focus
def focus6(event):
    # set focus on the phone_no_field box
    phone_no_field.focus_set()

# Function to set focus
def focus7(event):
    # set focus on the startDate_field box
    startDate_field.focus_set()

# Function to set focus
def focus8(event):
    # set focus on the endDate_field box
    endDate_field.focus_set()

# Function to set focus
def focus9(event):
    # set focus on the leaveReason_field box
    leaveReason_field.focus_set()
 
 
# Function for clearing the
# contents of text entry boxes
def clear():
     
    # clear the content of text entry box
    usn_field.delete(0, END)
    name_field.delete(0, END)
    branch_field.delete(0, END)
    semester_field.delete(0, END)
    section_field.delete(0, END)
    email_id_field.delete(0, END)
    phone_no_field.delete(0, END)
    startDate_field.delete(0, END)
    endDate_field.delete(0, END)
    leaveReason_field.delete(0, END)
 
 
# Function to take data from GUI
# window and write to an excel file
def insert():
     
    # if user not fill any entry
    # then print "empty input"
    if (usn_field.get() == "" and
        name_field.get() == "" and
        branch_field.get() == "" and
        semester_field.get() == "" and
        section_field.get() == "" and
        email_id_field.get() == "" and
        phone_no_field.get() == "" and
        startDate_field.get() == "" and
        endDate_field.get() == "" and
        leaveReason_field.get() == ""):
             
        print("empty input")
 
    else:
 
        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column
 
        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = usn_field.get()
        sheet.cell(row=current_row + 1, column=2).value = name_field.get()
        sheet.cell(row=current_row + 1, column=3).value = branch_field.get()
        sheet.cell(row=current_row + 1, column=4).value = semester_field.get()
        sheet.cell(row=current_row + 1, column=5).value = section_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = phone_no_field.get()
        sheet.cell(row=current_row + 1, column=8).value = startDate_field.get()
        sheet.cell(row=current_row + 1, column=9).value = endDate_field.get()
        sheet.cell(row=current_row + 1, column=10).value = leaveReason_field.get()
 
        # save the file
        wb.save('D:\\5th Sem Online Classes\\ADP\\ADP Activity\\excel.xlsx')
 
        # set focus on the name_field box
        usn_field.focus_set()
 
        # call the clear() function
        clear()
 
 
# Driver code
if __name__ == "__main__":
     
    # create a GUI window
    root = Tk()
 
    # set the background colour of GUI window
    root.configure(background='#AAEDFC')
 
    # set the title of GUI window
    root.title("Leave Application Form")
 
    # set the configuration of GUI window
    root.geometry("500x500")
 
    excel()
 
    # create a Form label
    heading = Label(root, text="Leave Application", bg="#AAEDFC")
 
    # create a USN label
    usn = Label(root, text="\nUSN", bg="#AAEDFC")
 
    # create a Name label
    name = Label(root, text="\nName", bg="#AAEDFC")
 
    # create a Branch label
    branch = Label(root, text="\nBranch", bg="#AAEDFC")
 
    # create a Semester label
    semester = Label(root, text="\nSemester", bg="#AAEDFC")
 
    # create a Section. label
    section = Label(root, text="\nSection", bg="#AAEDFC")
 
    # create a Email ID label
    email_id = Label(root, text="\nEmail ID", bg="#AAEDFC")
 
    # create a Phone No label
    phone_no = Label(root, text="\nPhone No", bg="#AAEDFC")

    # create a Start Date label
    startDate = Label(root, text="\nStart Date\n(dd/mm/yyyy)", bg="#AAEDFC")

    # create a End Date label
    endDate = Label(root, text="\nEnd Date\n(dd/mm/yyyy)", bg="#AAEDFC")

    # create a Start Date label
    leaveReason = Label(root, text="\nLeave Reason", bg="#AAEDFC")
 
    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    usn.grid(row=1, column=0)
    name.grid(row=2, column=0)
    branch.grid(row=3, column=0)
    semester.grid(row=4, column=0)
    section.grid(row=5, column=0)
    email_id.grid(row=6, column=0)
    phone_no.grid(row=7, column=0)
    startDate.grid(row=8, column=0)
    endDate.grid(row=9, column=0)
    leaveReason.grid(row=10, column=0)
 
    # create a text entry box
    # for typing the information
    usn_field = Entry(root)
    name_field = Entry(root)
    branch_field = Entry(root)
    semester_field = Entry(root)
    section_field = Entry(root)
    email_id_field = Entry(root)
    phone_no_field = Entry(root)
    startDate_field = Entry(root)
    endDate_field = Entry(root)
    leaveReason_field = Entry(root)
 
    # bind method of widget is used for
    # the binding the function with the events
 
    # whenever the enter key is pressed
    # then call the focus1 function
    usn_field.bind("<Return>", focus1)
 
    # whenever the enter key is pressed
    # then call the focus2 function
    name_field.bind("<Return>", focus2)
 
    # whenever the enter key is pressed
    # then call the focus3 function
    branch_field.bind("<Return>", focus3)
 
    # whenever the enter key is pressed
    # then call the focus4 function
    semester_field.bind("<Return>", focus4)
 
    # whenever the enter key is pressed
    # then call the focus5 function
    section_field.bind("<Return>", focus5)
 
    # whenever the enter key is pressed
    # then call the focus6 function
    email_id_field.bind("<Return>", focus6)

    # whenever the enter key is pressed
    # then call the focus7 function
    phone_no_field.bind("<Return>", focus7)

    # whenever the enter key is pressed
    # then call the focus8 function
    startDate_field.bind("<Return>", focus8)

    # whenever the enter key is pressed
    # then call the focus9 function
    endDate_field.bind("<Return>", focus9)


 
    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    usn_field.grid(row=1, column=1, ipadx="100")
    name_field.grid(row=2, column=1, ipadx="100")
    branch_field.grid(row=3, column=1, ipadx="100")
    semester_field.grid(row=4, column=1, ipadx="100")
    section_field.grid(row=5, column=1, ipadx="100")
    email_id_field.grid(row=6, column=1, ipadx="100")
    phone_no_field.grid(row=7, column=1, ipadx="100")
    startDate_field.grid(row=8, column=1, ipadx="100")
    endDate_field.grid(row=9, column=1, ipadx="100")
    leaveReason_field.grid(row=10, column=1, ipadx="100")
 
    # call excel function
    excel()
 
    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="White",
                            bg="#2B6673", command=insert)
    submit.grid(row=11, column=1)
 
    # start the GUI
    root.mainloop()